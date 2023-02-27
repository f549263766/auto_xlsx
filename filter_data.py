import argparse
import re
import os.path as osp
import pandas as pd
from openpyxl import load_workbook, Workbook
from itertools import zip_longest


def config_args():
    parser = argparse.ArgumentParser('time')
    parser.add_argument("--file_path", type=str,
                        default=r'D:\QYZ\Code\misc\result\output.xlsx',
                        help="root path of source data")
    parser.add_argument("--output_root", type=str,
                        default=r"D:\QYZ\Code\misc\result",
                        help="root path of source data")
    parsers = parser.parse_args()
    return parsers


def write_data(header_names, data_list, times):
    wb = Workbook()
    ws = wb.worksheets[0]
    ws.title = "过滤袋压差"
    header_names.insert(0, "时间")
    ws.append(header_names)
    data_list.insert(0, times)
    trans_data = [list(filter(None, i)) for i in zip_longest(*data_list)]

    for row_data in trans_data:
        ws.append(row_data)

    wb.save(osp.join(args.output_root, "过滤袋压差output.xlsx"))


def get_differential_pressure(table, key, sheet_name, header_names, data_list):
    try:
        differential_pressure = table.loc[:, key].values.tolist()
        if differential_pressure:
            header_names.append(f"{sheet_name}-{key}")
            data_list.append(differential_pressure)
    except KeyError:
        pass


def filter_data():
    # init store data list
    data_list = []
    # load excel with openpxl
    excel = load_workbook(args.file_path)
    # get list of all sheet names
    sheet_names = excel.sheetnames
    # close excel and release resources
    excel.close()

    # newly increased header name and time
    header_names, times = [], []
    # traverse the sheet name to find the target sheet
    for sheet_name in sheet_names:
        if re.search(r'过滤袋压差', sheet_name):
            table = pd.read_excel(args.file_path, sheet_name=sheet_name)
            get_differential_pressure(table, "上压差", sheet_name, header_names, data_list)
            get_differential_pressure(table, "下压差", sheet_name, header_names, data_list)
            times = table.loc[:, '时间'].values.tolist()

    write_data(header_names, data_list, times)


def main():
    filter_data()


if __name__ == '__main__':
    args = config_args()
    main()
